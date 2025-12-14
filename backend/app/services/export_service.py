"""
数据导出服务
Data Export Service

支持多种格式的数据导出：CSV、JSON、Excel。
支持异步后台导出大数据量。
"""

import asyncio
import csv
import io
import json
from datetime import datetime
from enum import Enum
from pathlib import Path
from typing import Any, Dict, List, Optional, Union

from sqlalchemy import select
from sqlalchemy.ext.asyncio import AsyncSession

from app.core.config import settings
from app.core.logging import get_logger
from app.models.export_task import ExportTask, ExportFormat, ExportStatus
from app.models.news_article import NewsArticle
from app.models.social_message import SocialMessage
from app.models.social_session import SocialSession
from app.services.storage_service import StorageService

logger = get_logger(__name__)


class DataSource(str, Enum):
    """导出数据源"""
    NEWS = "news"
    SOCIAL_SESSIONS = "social_sessions"
    SOCIAL_MESSAGES = "social_messages"


class ExportService:
    """
    数据导出服务

    支持功能：
    - 多格式导出（CSV、JSON、Excel）
    - 后台异步导出
    - 导出任务管理
    - 导出文件存储（本地/S3/MinIO）
    """

    # 每批处理的记录数
    BATCH_SIZE = 1000

    # 导出文件目录
    EXPORT_DIR = "exports"

    def __init__(
        self,
        db: AsyncSession,
        storage_service: Optional[StorageService] = None,
    ):
        """
        初始化导出服务

        Args:
            db: 数据库会话
            storage_service: 存储服务（可选）
        """
        self.db = db
        self.storage_service = storage_service
        self.logger = get_logger(__name__)

    async def create_export_task(
        self,
        data_source: DataSource,
        export_format: ExportFormat,
        filters: Optional[Dict[str, Any]] = None,
        filename: Optional[str] = None,
    ) -> ExportTask:
        """
        创建导出任务

        Args:
            data_source: 数据源类型
            export_format: 导出格式
            filters: 过滤条件
            filename: 自定义文件名

        Returns:
            创建的导出任务
        """
        # 生成文件名
        if not filename:
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            # 根据格式确定文件扩展名
            ext_map = {
                ExportFormat.CSV: "csv",
                ExportFormat.JSON: "json",
                ExportFormat.EXCEL: "xlsx",
            }
            ext = ext_map.get(export_format, "csv")
            filename = f"{data_source.value}_{timestamp}.{ext}"

        # 创建任务记录
        task = ExportTask(
            data_source=data_source.value,
            export_format=export_format.value,  # 使用枚举值
            filters=json.dumps(filters or {}),
            filename=filename,
            status=ExportStatus.PENDING.value,  # 使用枚举值
        )

        self.db.add(task)
        await self.db.commit()
        await self.db.refresh(task)

        self.logger.info(f"创建导出任务: {task.id}, 数据源: {data_source.value}")

        return task

    async def execute_export(self, task_id: int) -> ExportTask:
        """
        执行导出任务

        Args:
            task_id: 任务 ID

        Returns:
            更新后的任务
        """
        # 获取任务
        stmt = select(ExportTask).where(ExportTask.id == task_id)
        result = await self.db.execute(stmt)
        task = result.scalar_one_or_none()

        if not task:
            raise ValueError(f"导出任务不存在: {task_id}")

        # Check status (case-insensitive)
        allowed_statuses = [ExportStatus.PENDING.value.lower(), ExportStatus.FAILED.value.lower()]
        if task.status.lower() not in allowed_statuses:
            raise ValueError(f"任务状态不允许执行: {task.status}")

        # 更新状态为处理中
        task.status = ExportStatus.PROCESSING.value
        task.started_at = datetime.now()
        await self.db.commit()

        try:
            # 解析过滤条件
            filters = json.loads(task.filters) if task.filters else {}

            # 获取数据
            data_source = DataSource(task.data_source)
            data, total_count = await self._fetch_data(data_source, filters)

            task.total_records = total_count

            # 导出数据 (case-insensitive format comparison)
            export_format_upper = task.export_format.upper()
            if export_format_upper == ExportFormat.CSV.value:
                content = await self._export_to_csv(data, data_source)
                content_type = "text/csv"
            elif export_format_upper == ExportFormat.JSON.value:
                content = await self._export_to_json(data)
                content_type = "application/json"
            elif export_format_upper == ExportFormat.EXCEL.value:
                content = await self._export_to_excel(data, data_source)
                content_type = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
            else:
                raise ValueError(f"不支持的导出格式: {task.export_format}")

            # 保存文件
            file_path = await self._save_export_file(
                content=content,
                filename=task.filename,
                content_type=content_type,
            )

            # 更新任务状态
            task.status = ExportStatus.COMPLETED.value
            task.file_path = file_path
            task.file_size = len(content) if isinstance(content, bytes) else len(content.encode())
            task.completed_at = datetime.now()
            task.exported_records = total_count

            self.logger.info(
                f"导出完成: 任务 {task.id}, "
                f"记录数 {total_count}, "
                f"文件大小 {task.file_size} bytes"
            )

        except Exception as e:
            task.status = ExportStatus.FAILED.value
            task.error_message = str(e)
            self.logger.error(f"导出失败: 任务 {task.id}, 错误: {e}", exc_info=True)

        await self.db.commit()
        await self.db.refresh(task)

        return task

    async def _fetch_data(
        self,
        data_source: DataSource,
        filters: Dict[str, Any],
    ) -> tuple[List[Dict[str, Any]], int]:
        """
        获取导出数据

        Args:
            data_source: 数据源
            filters: 过滤条件

        Returns:
            (数据列表, 总记录数)
        """
        if data_source == DataSource.NEWS:
            return await self._fetch_news_data(filters)
        elif data_source == DataSource.SOCIAL_SESSIONS:
            return await self._fetch_social_sessions(filters)
        elif data_source == DataSource.SOCIAL_MESSAGES:
            return await self._fetch_social_messages(filters)
        else:
            raise ValueError(f"不支持的数据源: {data_source}")

    async def _fetch_news_data(
        self,
        filters: Dict[str, Any],
    ) -> tuple[List[Dict[str, Any]], int]:
        """获取新闻数据"""
        stmt = select(NewsArticle)

        # 应用过滤条件
        if filters.get("source_key"):
            stmt = stmt.where(NewsArticle.source_key == filters["source_key"])
        if filters.get("category"):
            stmt = stmt.where(NewsArticle.category == filters["category"])
        if filters.get("start_date"):
            start = datetime.fromisoformat(filters["start_date"])
            stmt = stmt.where(NewsArticle.published_at >= start)
        if filters.get("end_date"):
            end = datetime.fromisoformat(filters["end_date"])
            stmt = stmt.where(NewsArticle.published_at <= end)

        stmt = stmt.order_by(NewsArticle.published_at.desc())

        result = await self.db.execute(stmt)
        articles = result.scalars().all()

        data = []
        for article in articles:
            data.append({
                "id": article.id,
                "title": article.title,
                "url": article.url,
                "source_key": article.source_key,
                "category": article.category,
                "published_at": article.published_at.isoformat() if article.published_at else None,
                "created_at": article.created_at.isoformat() if article.created_at else None,
            })

        return data, len(data)

    async def _fetch_social_sessions(
        self,
        filters: Dict[str, Any],
    ) -> tuple[List[Dict[str, Any]], int]:
        """获取社交会话数据"""
        stmt = select(SocialSession)

        if filters.get("platform"):
            stmt = stmt.where(SocialSession.platform == filters["platform"])
        if filters.get("status"):
            stmt = stmt.where(SocialSession.status == filters["status"])

        stmt = stmt.order_by(SocialSession.created_at.desc())

        result = await self.db.execute(stmt)
        sessions = result.scalars().all()

        data = []
        for session in sessions:
            data.append({
                "id": session.id,
                "session_key": session.session_key,
                "platform": session.platform.value,
                "target_type": session.target_type.value,
                "target_id": session.target_id,
                "target_name": session.target_name,
                "status": session.status.value,
                "message_count": session.message_count,
                "last_message_at": session.last_message_at.isoformat() if session.last_message_at else None,
                "created_at": session.created_at.isoformat() if session.created_at else None,
            })

        return data, len(data)

    async def _fetch_social_messages(
        self,
        filters: Dict[str, Any],
    ) -> tuple[List[Dict[str, Any]], int]:
        """获取社交消息数据"""
        stmt = select(SocialMessage)

        if filters.get("session_id"):
            stmt = stmt.where(SocialMessage.session_id == filters["session_id"])
        if filters.get("start_date"):
            start = datetime.fromisoformat(filters["start_date"])
            stmt = stmt.where(SocialMessage.posted_at >= start)
        if filters.get("end_date"):
            end = datetime.fromisoformat(filters["end_date"])
            stmt = stmt.where(SocialMessage.posted_at <= end)

        stmt = stmt.order_by(SocialMessage.posted_at.desc())

        result = await self.db.execute(stmt)
        messages = result.scalars().all()

        data = []
        for msg in messages:
            data.append({
                "id": msg.id,
                "session_id": msg.session_id,
                "message_id": msg.message_id,
                "author_id": msg.author_id,
                "author_name": msg.author_name,
                "author_username": msg.author_username,
                "content": msg.content,
                "reply_count": msg.reply_count,
                "repost_count": msg.repost_count,
                "like_count": msg.like_count,
                "view_count": msg.view_count,
                "posted_at": msg.posted_at.isoformat() if msg.posted_at else None,
                "created_at": msg.created_at.isoformat() if msg.created_at else None,
            })

        return data, len(data)

    async def _export_to_csv(
        self,
        data: List[Dict[str, Any]],
        data_source: DataSource,
    ) -> str:
        """导出为 CSV 格式"""
        if not data:
            return ""

        output = io.StringIO()
        writer = csv.DictWriter(output, fieldnames=data[0].keys())
        writer.writeheader()
        writer.writerows(data)

        return output.getvalue()

    async def _export_to_json(
        self,
        data: List[Dict[str, Any]],
    ) -> str:
        """导出为 JSON 格式"""
        return json.dumps(data, ensure_ascii=False, indent=2)

    async def _export_to_excel(
        self,
        data: List[Dict[str, Any]],
        data_source: DataSource,
    ) -> bytes:
        """导出为 Excel 格式"""
        try:
            import openpyxl
            from openpyxl import Workbook
        except ImportError:
            raise ImportError("需要安装 openpyxl 库才能导出 Excel 格式")

        wb = Workbook()
        ws = wb.active
        ws.title = data_source.value

        if not data:
            return self._workbook_to_bytes(wb)

        # 写入表头
        headers = list(data[0].keys())
        for col, header in enumerate(headers, 1):
            ws.cell(row=1, column=col, value=header)

        # 写入数据
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                ws.cell(row=row_idx, column=col_idx, value=row_data.get(header))

        return self._workbook_to_bytes(wb)

    def _workbook_to_bytes(self, workbook) -> bytes:
        """将 Workbook 转换为字节"""
        output = io.BytesIO()
        workbook.save(output)
        output.seek(0)
        return output.read()

    async def _save_export_file(
        self,
        content: Union[str, bytes],
        filename: str,
        content_type: str,
    ) -> str:
        """
        保存导出文件

        Args:
            content: 文件内容
            filename: 文件名
            content_type: MIME 类型

        Returns:
            文件路径或 URL
        """
        # 确保是字节类型
        if isinstance(content, str):
            content_bytes = content.encode("utf-8")
        else:
            content_bytes = content

        # 使用存储服务
        if self.storage_service:
            file_path = f"{self.EXPORT_DIR}/{filename}"
            await self.storage_service.upload(
                file_path=file_path,
                content=content_bytes,
                content_type=content_type,
            )
            return file_path

        # 本地存储
        export_dir = Path(settings.DATA_DIR) / self.EXPORT_DIR
        export_dir.mkdir(parents=True, exist_ok=True)

        file_path = export_dir / filename
        file_path.write_bytes(content_bytes)

        return str(file_path)

    async def get_download_url(self, task: ExportTask) -> Optional[str]:
        """
        获取导出文件下载 URL

        Args:
            task: 导出任务

        Returns:
            下载 URL 或本地路径
        """
        if not task.file_path:
            return None

        if self.storage_service:
            return await self.storage_service.get_url(task.file_path)

        return task.file_path

    async def cleanup_old_exports(self, days: int = 7) -> int:
        """
        清理过期的导出文件

        Args:
            days: 保留天数

        Returns:
            清理的任务数
        """
        from datetime import timedelta

        cutoff = datetime.now() - timedelta(days=days)

        stmt = select(ExportTask).where(
            ExportTask.created_at < cutoff,
            ExportTask.status == ExportStatus.COMPLETED.value,
        )

        result = await self.db.execute(stmt)
        tasks = result.scalars().all()

        cleaned = 0
        for task in tasks:
            try:
                if task.file_path:
                    if self.storage_service:
                        await self.storage_service.delete(task.file_path)
                    else:
                        file_path = Path(task.file_path)
                        if file_path.exists():
                            file_path.unlink()

                await self.db.delete(task)
                cleaned += 1

            except Exception as e:
                self.logger.error(f"清理导出任务失败: {task.id}, 错误: {e}")

        await self.db.commit()

        self.logger.info(f"清理过期导出: {cleaned} 个任务")

        return cleaned
